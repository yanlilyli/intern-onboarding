from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

wb = Workbook()

# ============ Sheet 1: 6个月工作计划表 ============
ws = wb.active
ws.title = "6个月工作计划"

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
phase_fills = [
    PatternFill('solid', fgColor='F2F7FC'),
    PatternFill('solid', fgColor='FFF8E1'),
    PatternFill('solid', fgColor='F1F8E9'),
    PatternFill('solid', fgColor='FBE9E7'),
    PatternFill('solid', fgColor='F3E5F5'),
]

# 标题
ws.merge_cells('A1:F1')
ws['A1'] = '实习生6个月工作计划 — 招聘开源方向（美术开源 + 策划Sourcing&Mapping）'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('A2:F2')
ws['A2'] = '实习生：[姓名]  |  Leader：李彦  |  入职日期：2026年5月  |  计划周期：2026.05 - 2026.11（6个月）'
ws['A2'].font = Font(name='Arial', size=10, color='666666')
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

col_widths = [14, 14, 28, 42, 24, 22]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

headers = ['阶段', '时间', '工作目标', '具体任务', '产出/交付物', '备注']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# ====== 6个月计划数据 ======
plan_data = [
    # 第一阶段：快速上手 (W1-W2) —— 已有招聘实习经验，直接sourcing实操
    {'phase': '第一阶段\n快速上手\n(W1-W2)', 'rows': [
        ['W1\n第1周',
         '业务熟悉 + sourcing实操启动',
         '1. 了解S工作室/天玑组织架构、在招方向和人才画像\n2. 学习Offer系统操作，配置Boss/猎聘账号\n3. 直接开始美术方向sourcing实操（已有招聘经验，快速上手）\n4. 掌握美术方向搜索关键词体系和渠道特点\n5. 了解竞品工作室（人才都在哪里）',
         '首批简历推荐\n搜索关键词库v1\n业务学习笔记',
         '有招聘经验可以W1\n就开始出活'],
        ['W2\n第2周',
         'sourcing放量 + mapping方法学习',
         '1. 美术sourcing持续推进，熟悉各渠道操作\n2. 学习mapping方法论（组织架构拆解法）\n3. 跟听1-2场面试，建立岗位判断力\n4. 和业务方对齐策划岗位人才画像\n5. 初步锁定策划mapping目标公司',
         '美术简历持续推荐\n策划画像对齐记录',
         'W2开始为策划方向做准备'],
    ]},

    # 第二阶段：双线并行 (W3-W8)
    {'phase': '第二阶段\n双线并行\n(W3-W8)', 'rows': [
        ['W3-W4\n第3-4周',
         '美术开源双线放量\n（S校招 + 天玑社招）',
         '1. S美术泛校招sourcing放量（原画/3D/TA/场景等）\n2. 天玑美术社招sourcing同步放量（ArtStation/脉脉/LinkedIn）\n3. 两个方向本质相同——都是美术开源，校招走院校pool，社招走行业渠道\n4. 建立候选人标签体系和跟进状态表\n5. 打磨触达话术，提升回复率',
         '美术候选人pool表\n各渠道简历推荐\n周报',
         '校招/社招是同一件事\n只是渠道风格不同'],
        ['W5-W6\n第5-6周',
         '策划sourcing启动\nmapping辅助产出',
         '1. S新项目资深策划岗位sourcing正式启动\n2. 通过mapping锁定目标公司+核心人选名单\n3. 策划方向开始定向触达和沟通\n4. 美术双线sourcing持续稳定产出\n5. mapping是手段，sourcing交付简历才是目的',
         '策划方向简历推荐\nMapping报告v1\n美术周报',
         'Mapping帮助精准sourcing\n但交付是核心衡量'],
        ['W7-W8\n第7-8周',
         '策划sourcing深入\n美术渠道优化',
         '1. 策划sourcing扩展到3-5家目标公司的核心人选\n2. 深度拆解目标公司策划团队架构，持续补充人选名单\n3. 主动推进策划候选人进面试流程\n4. 美术sourcing分析渠道效果，聚焦高效渠道\n5. 推动优质美术候选人进面试',
         '策划简历持续推荐\nMapping报告v2\n渠道效果分析',
         '从量转向质\n关注进面转化率'],
    ]},

    # 第三阶段：稳定产出 (W9-W14)
    {'phase': '第三阶段\n稳定产出\n(W9-W14)', 'rows': [
        ['W9-W10\n第9-10周',
         '双线稳定交付\n推动面试转化',
         '1. 美术+策划双线sourcing稳定产出\n2. 策划mapping报告完善（组织架构/人才分布/触达策略/竞争分析）\n3. 重点跟进高潜候选人，推动面试安排\n4. 候选人意愿度管理和跟进\n5. 复盘前期数据，优化sourcing策略',
         '双线简历持续交付\nMapping完整报告\n面试跟进表',
         '重点从开源转向转化\n推动结果落地'],
        ['W11-W12\n第11-12周',
         '转化攻坚\n深度候选人运营',
         '1. 策划+美术高潜候选人重点推进面试\n2. 面试安排和全程协调\n3. 二次激活此前未回复的优质候选人\n4. 策划mapping持续补充新发现的人选\n5. 和业务方保持画像同步，及时调整方向',
         '面试推进记录\n候选人运营SOP\n周报',
         '这个阶段盯结果\n面试通过率是关键'],
        ['W13-W14\n第13-14周',
         '方法论沉淀\n查漏补缺',
         '1. 整理sourcing方法论文档（渠道选择/关键词/话术/mapping流程）\n2. 更新mapping报告终稿\n3. 查漏补缺：未完成的方向补量\n4. 总结各渠道投入产出比',
         '方法论文档v1\nMapping终稿\n渠道ROI分析',
         '沉淀可复用的知识'],
    ]},

    # 第四阶段：独立运作 (W15-W20)
    {'phase': '第四阶段\n独立运作\n(W15-W20)', 'rows': [
        ['W15-W16\n第15-16周',
         '独立承担sourcing模块',
         '1. 独立负责美术/策划sourcing的周度规划和执行\n2. 减少leader干预，自主决策搜索方向和资源分配\n3. 尝试承接新岗位方向的sourcing\n4. 如有新实习生入职，协助带教sourcing基础',
         '独立周报\n新方向候选人推荐',
         '进入独立工作状态\n自己规划自己执行'],
        ['W17-W18\n第17-18周',
         '深化行业认知\n持续产出',
         '1. 持续双线sourcing稳定产出\n2. 建立对游戏行业人才市场的深度认知\n3. 积累行业人脉和信息网络\n4. 输出人才市场洞察（竞品动态/人才流动趋势）',
         '人才市场洞察报告\n持续简历交付',
         '培养真正的招聘sense'],
        ['W19-W20\n第19-20周',
         '成果巩固\n为收尾做准备',
         '1. 各方向sourcing成果盘点\n2. 候选人资料和跟进状态全量整理\n3. 知识库和文档归档\n4. 梳理待交接事项清单',
         '成果盘点表\n知识库整理',
         '开始为交接做准备'],
    ]},

    # 第五阶段：收尾交接 (W21-W24)
    {'phase': '第五阶段\n收尾交接\n(W21-W24)', 'rows': [
        ['W21-W22\n第21-22周',
         '知识交接\n经验沉淀',
         '1. 整理所有候选人资料，确保状态可追溯、可交接\n2. 输出完整实习总结（方法论+数据成果+改进建议）\n3. 制作团队知识交接文档\n4. 带教交接给后续同事',
         '交接文档\n实习总结报告',
         '确保工作无缝衔接\n不留黑洞'],
        ['W23-W24\n第23-24周',
         '复盘总结\n实习答辩',
         '1. 完成实习答辩PPT\n2. 向团队做实习成果分享\n3. 收集leader和同事反馈\n4. 职业规划沟通',
         '答辩PPT\n成果分享',
         '争取好评/转正\n保持长期联系'],
    ]},
]

# 填充数据
row = 5
for phase_idx, phase in enumerate(plan_data):
    phase_start = row
    for r_data in phase['rows']:
        ws.cell(row=row, column=2, value=r_data[0]).font = bold_font
        ws.cell(row=row, column=2).alignment = center_align
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=r_data[1]).font = normal_font
        ws.cell(row=row, column=3).alignment = left_align
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value=r_data[2]).font = normal_font
        ws.cell(row=row, column=4).alignment = left_align
        ws.cell(row=row, column=4).border = thin_border

        ws.cell(row=row, column=5, value=r_data[3]).font = normal_font
        ws.cell(row=row, column=5).alignment = left_align
        ws.cell(row=row, column=5).border = thin_border

        ws.cell(row=row, column=6, value=r_data[4]).font = normal_font
        ws.cell(row=row, column=6).alignment = left_align
        ws.cell(row=row, column=6).border = thin_border

        ws.row_dimensions[row].height = 100
        row += 1

    phase_end = row - 1
    ws.cell(row=phase_start, column=1, value=phase['phase']).font = bold_font
    ws.cell(row=phase_start, column=1).alignment = center_align
    ws.cell(row=phase_start, column=1).fill = phase_fills[phase_idx % len(phase_fills)]
    ws.cell(row=phase_start, column=1).border = thin_border
    if phase_start != phase_end:
        ws.merge_cells(f'A{phase_start}:A{phase_end}')
    for r in range(phase_start + 1, phase_end + 1):
        ws.cell(row=r, column=1).border = thin_border

ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 22

# ============ Sheet 2: 周度跟踪表 (24周) ============
ws2 = wb.create_sheet("周度跟踪")

ws2.merge_cells('A1:H1')
ws2['A1'] = '实习生周度工作跟踪表（6个月/24周）'
ws2['A1'].font = Font(name='Arial', bold=True, size=13, color='2F5496')
ws2['A1'].alignment = center_align

headers2 = ['周次', '日期范围', '本周核心任务', '完成情况', '美术推荐数', '策划推荐数', '进面人数', 'Leader反馈']
col_widths2 = [8, 16, 32, 32, 12, 12, 10, 28]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

start_date = datetime.date(2026, 5, 12)
for week in range(1, 25):
    r = week + 3
    ws2.cell(row=r, column=1, value=f'W{week}').font = bold_font
    ws2.cell(row=r, column=1).alignment = center_align
    ws2.cell(row=r, column=1).border = thin_border

    week_start = start_date + datetime.timedelta(weeks=week-1)
    week_end = week_start + datetime.timedelta(days=4)
    ws2.cell(row=r, column=2, value=f'{week_start.strftime("%m/%d")}-{week_end.strftime("%m/%d")}')
    ws2.cell(row=r, column=2).alignment = center_align
    ws2.cell(row=r, column=2).border = thin_border

    for col in range(3, 9):
        ws2.cell(row=r, column=col).border = thin_border
        ws2.cell(row=r, column=col).alignment = left_align if col <= 4 else center_align

    ws2.row_dimensions[r].height = 30

# ============ Sheet 3: 工作目标总览 ============
ws3 = wb.create_sheet("工作目标总览")

ws3.merge_cells('A1:D1')
ws3['A1'] = '实习期工作目标总览'
ws3['A1'].font = Font(name='Arial', bold=True, size=13, color='2F5496')
ws3['A1'].alignment = center_align

headers3 = ['工作目标', '覆盖阶段', '具体说明', '当前进展']
col_widths3 = [30, 15, 45, 20]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

goal_data = [
    ['快速上手sourcing全流程', '第一阶段', '利用已有招聘经验，W1即开始sourcing实操产出', ''],
    ['S美术泛校招开源', '第一~四阶段', '覆盖原画/3D/TA/场景等方向，通过校园渠道建立美术人才pool', ''],
    ['天玑美术社招开源', '第一~四阶段', '通过ArtStation/脉脉/LinkedIn等行业渠道sourcing社招美术人才', ''],
    ['S新项目资深策划sourcing', '第二~四阶段', '通过mapping精准锁定目标人选，sourcing交付简历是核心衡量', ''],
    ['策划方向mapping报告', '第二~三阶段', 'mapping辅助sourcing，产出组织架构+人才名单+触达策略', ''],
    ['推动候选人面试转化', '第二~四阶段', '不止于推简历，跟进到进面试、拿结果', ''],
    ['沉淀sourcing方法论', '第三~四阶段', '输出可复用的渠道/关键词/话术/mapping方法论文档', ''],
    ['独立承担sourcing模块', '第四阶段', '减少leader干预，自主规划和执行sourcing任务', ''],
    ['完成知识交接与复盘', '第五阶段', '确保工作可追溯、可交接，输出实习总结', ''],
]

for i, data in enumerate(goal_data, 4):
    for col, val in enumerate(data, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        cell.font = normal_font
        cell.alignment = left_align if col >= 3 else center_align
        cell.border = thin_border
    ws3.row_dimensions[i].height = 28

# 保存
output_path = '/Users/liyan/WorkBuddy/20260507124820/output/intern_6month_plan.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")
